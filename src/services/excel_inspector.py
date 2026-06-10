"""Fast, read-only inspection of an uploaded survey Excel.

Reads only the header row plus a small sample (for value previews + filter
detection) without loading the whole sheet.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from config import settings
from models.upload import ColumnInfo, DetectedFilters, UploadResponse

_SAMPLE_ROWS = 20
_MAX_SAMPLE_VALUES = 5


def upload_path(upload_id: str) -> Path:
    return settings.uploads_dir / f"{upload_id}.xlsx"


def _norm(name: object) -> str:
    return str(name).strip().lower() if name is not None else ""


def _detect_column(headers: list[str], *exact: str) -> tuple[int | None, str | None]:
    """Find a column by exact normalized match first, then substring."""
    norm = [_norm(h) for h in headers]
    for target in exact:
        for i, h in enumerate(norm):
            if h == target:
                return i, headers[i]
    for target in exact:
        for i, h in enumerate(norm):
            if target in h:
                return i, headers[i]
    return None, None


def inspect_excel(path: str | Path) -> UploadResponse:
    """Inspect an Excel file's first sheet: headers, sample values, detected filters."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        sheet_name = ws.title

        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            raise ValueError("The Excel file has no header row")

        headers = [(str(h).strip() if h is not None else f"column_{i+1}") for i, h in enumerate(header_row)]

        # Collect a small sample for value previews.
        samples: list[set[str]] = [set() for _ in headers]
        for n, row in enumerate(row_iter):
            if n >= _SAMPLE_ROWS:
                break
            for i, cell in enumerate(row):
                if i < len(headers) and cell not in (None, ""):
                    if len(samples[i]) < _MAX_SAMPLE_VALUES:
                        samples[i].add(str(cell))

        status_idx, status_name = _detect_column(headers, "status")
        exclude_idx, exclude_name = _detect_column(headers, "exclude")
        filter_idxs = {idx for idx in (status_idx, exclude_idx) if idx is not None}

        columns = [
            ColumnInfo(
                name=h,
                index=i,
                sample_values=sorted(samples[i]),
                is_filter=i in filter_idxs,
            )
            for i, h in enumerate(headers)
        ]
        candidate_labels = [h for i, h in enumerate(headers) if i not in filter_idxs]

        try:
            estimated_rows = ws.max_row
        except Exception:
            estimated_rows = None

        warnings: list[str] = []
        if status_name is None:
            warnings.append("No 'status' column found — the agent will use all rows (no submitted filter).")
        if exclude_name is None:
            warnings.append("No 'exclude' column found — no respondents will be dropped by exclusion.")

        return UploadResponse(
            upload_id="",  # set by the route after saving
            sheet_name=sheet_name,
            columns=columns,
            candidate_labels=candidate_labels,
            filters=DetectedFilters(
                status_column=status_name,
                exclude_column=exclude_name,
                status_detected=status_name is not None,
                exclude_detected=exclude_name is not None,
                estimated_rows=estimated_rows,
            ),
            warnings=warnings,
        )
    finally:
        wb.close()


def column_names(path: str | Path) -> list[str]:
    """Header names only (used to validate a run's segment_by choices)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        try:
            header_row = next(ws.iter_rows(values_only=True))
        except StopIteration:
            return []
        return [(str(h).strip() if h is not None else f"column_{i+1}") for i, h in enumerate(header_row)]
    finally:
        wb.close()
