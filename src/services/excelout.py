"""Build the raw-data Excel workbook for a generate-all run.

Layout (sheet "Raw Data"): one row per respondent —
  respondent_id | status | submitDate | <one col per REAL survey question> | <one col per GENERATED question>
Real columns come from the survey's canonical question list (stable order across rows);
generated columns are the questions the researcher chose. A second "Errors" sheet is
added only if some respondents failed mid-run.
"""
from __future__ import annotations

import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from services import data

EXPORTS_DIR = Path(__file__).resolve().parents[2] / "exports"


def _join_real(entry: dict) -> str:
    vals = [a.get("answer") or a.get("value") or "" for a in (entry.get("answers") or [])]
    return ", ".join(v for v in vals if v)


def build_workbook(survey_id: str, gen_questions: list[dict], results: list[tuple]) -> tuple[Path, str]:
    """results: list of (respondent_doc, generated_answers|[], error|None)."""
    qmeta = data.get_survey_questions(survey_id)
    real_cols = [(q.sqlQuestionId, (q.text or q.label or f"Q{q.sqlQuestionId}")) for q in qmeta]
    gen_cols = [(q["id"], q.get("text") or f"AI Q{q['id']}") for q in gen_questions]

    header = (
        ["respondent_id", "status", "submitDate"]
        + [h for _, h in real_cols]
        + [f"[AI] {h}" for _, h in gen_cols]
    )
    real_idx = {qid: 3 + i for i, (qid, _) in enumerate(real_cols)}
    gen_idx = {gid: 3 + len(real_cols) + i for i, (gid, _) in enumerate(gen_cols)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    ws.append(header)

    for doc, gen_answers, _err in results:
        row = [""] * len(header)
        row[0] = str(doc.get("_id", ""))
        row[1] = doc.get("status") or ""
        row[2] = data.jsonable(doc.get("submitDate")) or ""
        for qid, entry in data.respondent_answers_by_qid(doc).items():
            j = real_idx.get(qid)
            if j is not None:
                row[j] = _join_real(entry)
        for a in (gen_answers or []):
            j = gen_idx.get(a.get("id"))
            if j is not None:
                row[j] = ", ".join(a.get("answer") or [])
        ws.append(row)

    ws.freeze_panes = "A2"
    # reasonable, capped column widths
    for col in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(40, max(12, len(str(header[col - 1])) + 2))

    errors = [(doc, err) for doc, _ga, err in results if err]
    if errors:
        es = wb.create_sheet("Errors")
        es.append(["respondent_id", "error"])
        for doc, err in errors:
            es.append([str(doc.get("_id", "")), err])

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    fname = f"raw_data_{survey_id}_{int(time.time())}.xlsx"
    path = EXPORTS_DIR / fname
    wb.save(path)
    return path, fname


def build_backtest_workbook(survey_id: str, seed_qids: set[int], holdout_qs: list[dict],
                            results: list[tuple]) -> tuple[Path, str]:
    """Backtest export. results: list of (respondent_doc, predicted_answers|[], error|None).

    Sheet "Backtest": one row per respondent —
      respondent_id | accuracy | <Seed: Q>… | per held-out Q: <Real: Q> | <[AI] Q> | <match: Q>
    """
    from services.backtest_service import compare_answer, option_aliases  # local import avoids a cycle

    qmeta = {q.sqlQuestionId: q for q in data.get_survey_questions(survey_id)}
    aliases_by_qid = {q["id"]: option_aliases(qmeta.get(q["id"])) for q in holdout_qs}
    seed_cols = [(qid, (qmeta[qid].text or qmeta[qid].label or f"Q{qid}"))
                 for qid in seed_qids if qid in qmeta]
    hold_cols = [(q["id"], q.get("text") or f"Q{q['id']}", q["type"]) for q in holdout_qs]

    header = ["respondent_id", "accuracy"]
    header += [f"Seed: {t}" for _, t in seed_cols]
    for _, t, _ty in hold_cols:
        header += [f"Real: {t}", f"[AI] {t}", f"match: {t}"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest"
    ws.append(header)

    def real_vals(answered: dict, qid: int) -> list[str]:
        entry = answered.get(qid)
        if not entry:
            return []
        vals = [a.get("answer") or a.get("value") or "" for a in (entry.get("answers") or [])]
        return [v for v in vals if v]

    for doc, pred_answers, _err in results:
        answered = data.respondent_answers_by_qid(doc)
        pred_by_id = {a.get("id"): (a.get("answer") or []) for a in (pred_answers or [])}
        row = [str(doc.get("_id", "")), ""]
        for qid, _t in seed_cols:
            row.append(", ".join(real_vals(answered, qid)))

        matched = scored = 0
        for qid, _t, qtype in hold_cols:
            actual = real_vals(answered, qid)
            predicted = pred_by_id.get(qid, [])
            if actual:
                cmp = compare_answer(qtype, predicted, actual, aliases_by_qid.get(qid))
                if cmp["scored"]:
                    scored += 1
                    if cmp["match"]:
                        matched += 1
                        mark = "✓"
                    else:
                        mark = "✗"
                else:
                    mark = "—"  # open-ended, not scored
            else:
                mark = ""  # respondent never answered → no ground truth
            row += [", ".join(actual), ", ".join(predicted), mark]
        row[1] = f"{round(100 * matched / scored)}%" if scored else ""
        ws.append(row)

    ws.freeze_panes = "C2"
    for col in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(40, max(12, len(str(header[col - 1])) + 2))

    errors = [(doc, err) for doc, _pa, err in results if err]
    if errors:
        es = wb.create_sheet("Errors")
        es.append(["respondent_id", "error"])
        for doc, err in errors:
            es.append([str(doc.get("_id", "")), err])

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    fname = f"backtest_{survey_id}_{int(time.time())}.xlsx"
    path = EXPORTS_DIR / fname
    wb.save(path)
    return path, fname
